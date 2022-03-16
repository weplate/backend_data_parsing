import datetime
import itertools
import re
from collections import defaultdict

import pytz
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.models import MealItem, MealSelection, School
from data.babson.add_items import SCHOOL_ID
from data.babson.common import MEAL_TIMES_DICT, MEAL_SELECTION_COLS


# input is a list of tuples (meal_item, meal, cafeteria_id) (meal = 'breakfast', 'lunch', 'dinner')
def add_meals(meal_items: list[tuple[MealItem, str, str]], menu_path, version):
    # setup item dicts from meal items input
    items_by_id: dict[str, MealItem] = {item.cafeteria_id: item for item, _, __ in meal_items}
    meal_sheet: Worksheet = load_workbook(menu_path, read_only=True)['Report']

    # dicts
    item_ids_by_day: dict[datetime.date, list[tuple[MealItem, str]]] = defaultdict(list)
    meals: list[MealSelection] = []
    corr_items: list[list[MealItem]] = []
    items_to_update: list[MealItem] = []
    babson = School.objects.get(pk=SCHOOL_ID)

    for col in MEAL_SELECTION_COLS:
        meal_day = None
        item_name = None
        meal = None
        for row in meal_sheet.iter_rows(13, meal_sheet.max_row):

            if row[0].value and (m := re.search(r'((Breakfast)|(Lunch)|(Dinner)) : ', row[0].value)):
                meal = m.group(1).lower()

            val = row[col].value
            # if meal_day == datetime.date(year=2022, month=3, day=7):
            #     print(val, val in items_by_id)
            if val is None:
                continue
            elif m := re.match(r'\w+ \((\d+)/(\d+)/(\d+)\)', val):
                meal_day = datetime.date(year=int(m.group(3)), month=int(m.group(1)), day=int(m.group(2)))
            elif val in items_by_id:
                item = items_by_id[val]
                item_ids_by_day[meal_day].append((item, meal))
                if item.station in ['FLAME', '500 DEGREES', 'CARVED AND CRAFTED']:
                    item_ids_by_day[meal_day].append((item, 'afterlunch'))
                item.name = item_name
                items_to_update.append(item)
            else:
                item_name = val
    print('Parsed spreadsheet')

    # Create objects
    for day, item_ids in item_ids_by_day.items():
        # _i = itertools.chain(*(items_by_id[item_id] for item_id in item_ids))
        _snd = lambda t: t[1]

        for meal, items in itertools.groupby(sorted(item_ids, key=_snd), key=_snd):
            # The Waterloo student did this one
            dt = datetime.datetime.combine(day, MEAL_TIMES_DICT[meal], pytz.timezone('America/Toronto'))
            m = MealSelection(name=f'{meal.title()} on {day.strftime("%A, %B")} {day.day}, {day.year}',
                              group=meal,
                              timestamp=dt,
                              school=babson,
                              version=version)
            meals.append(m)
            corr_items.append(list(map(lambda x: x[0], items)))
    print('Created Meal Selection Objects')

    meals = MealSelection.objects.bulk_create(meals)
    print(f'Created {len(meals)} meal objects')
    for idx, (m, i) in enumerate(zip(meals, corr_items), start=1):
        m.items.set(i)
        m.save()

        if idx % 40 == 0:
            print(f'Updated `items` field for {idx} meal objects')

    MealItem.objects.bulk_update(items_to_update, fields=('name',))
    print(f'Renamed {len(items_to_update)} items')
