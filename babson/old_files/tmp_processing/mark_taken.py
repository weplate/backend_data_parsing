import json
import pathlib

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

FILE_DIR = pathlib.Path(__file__).resolve().parent

RED_FILL = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')


def main():
    wb = load_workbook(FILE_DIR / 'MenuWorks_FDA_Menu_Main.xlsx')
    sheet: Worksheet = wb['Report']

    with open(FILE_DIR / 'good.json') as f:
        ids = set((x.strip() for x in json.load(f)))

    cnt = 0
    for row in sheet.rows:
        if row[1].value and row[1].value.strip() in ids:
            cnt += 1
            row[1].fill = RED_FILL

    sheet['A1'].fill = RED_FILL

    print(f'Found {cnt} ids')

    wb.save(FILE_DIR / 'MenuWorks_FDA_Menu_Main.xlsx')


if __name__ == '__main__':
    main()
