import json
import pathlib

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

FILE_DIR = pathlib.Path(__file__).resolve().parent


def main():
    wb = load_workbook(FILE_DIR / 'MenuWorks_FDA_Menu_Main_serving_volume__categories.xlsx')
    write_sheet: Worksheet = wb['Report']
    read_sheet: Worksheet = load_workbook(FILE_DIR / 'Nutrition_Facts_as_of_Feb_20_1.xlsx')['Report']

    with open(FILE_DIR / 'good.json') as f:
        ids = set((x.strip() for x in json.load(f)))

    id_mp = {}

    for row in read_sheet.rows:
        if row[2].value in ids:
            id_mp[row[2].value] = {
                'category': row[1].value,
                'portion_size': row[4].value
            }

    cnt = 0
    for row in write_sheet.rows:
        if row[1].value in ids:
            d = id_mp[row[1].value]
            row[3].value = d['category']
            row[4].value = d['portion_size']
            cnt += 1

    print(f'Updated {cnt} ids')

    wb.save(FILE_DIR / 'MenuWorks_FDA_Menu_Main.xlsx')


if __name__ == '__main__':
    main()
