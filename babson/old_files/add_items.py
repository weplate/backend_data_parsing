import json
import pathlib
import re
from re import Match
from typing import Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import serializers

from backend.models import MealItem

OUT_FILE_PATH = 'meal_items.json'
SCHOOL_ID = 10
MAX_NAME_LEN = 64
FILE_DIR = pathlib.Path(__file__).resolve().parent


def parse_meal_items(version):
    main_sheet: Worksheet = load_workbook(FILE_DIR / 'MenuWorks_FDA_Menu_Main.xlsx', read_only=True)['Report']
    micros_sheet: Worksheet = load_workbook(FILE_DIR / 'MenuWorks_FDA_Menu_Alt.xlsx', read_only=True)['Report']

    meal_items = {}

    NUM = r'(([0-9]+)(\/[0-9]+)?)'

    def parse_num(matches: Match):
        mat = matches.group(1)
        if '/' in mat:
            n, d = mat.split('/')
            return float(n) / float(d)
        else:
            return float(mat)

    bad_units = set()

    def parse_portion(value):
        if m := re.search(rf'{NUM} [cx]up', value):
            return parse_num(m) * 236.588, False
        elif m := re.search(rf'{NUM} pint', value):
            return parse_num(m) * 473, False
        elif m := re.search(rf'{NUM} tbsp', value):
            return parse_num(m) * 14.7866, False
        elif m := re.search(rf'{NUM} tsp', value):
            return parse_num(m) * 4.92892, False
        elif m := re.search(rf'{NUM} floz', value):
            return parse_num(m) * 29.5735, False
        elif m := re.search(rf'[0-9] ladle-{NUM}oz', value):
            return parse_num(m) * 29.5735, False
        else:
            bad_units.add(value)
            return -1, True

    # Read macros
    cur_station = None
    unparseable = 0
    for row in main_sheet.iter_rows(13):
        def get_col(col: Union[int, str]):
            if isinstance(col, int):
                return row[col].value
            elif isinstance(col, str) and len(col) == 1:
                return row[ord(col) - 65].value
            else:
                raise ValueError(f'Unreadable column {col}')

        def num_col(col: Union[int, str]):
            val = get_col(col)
            if val is None:
                return 0
            else:
                only_num = re.sub(r'[^0-9.]', '', val)
                return float(only_num) if only_num else 0

        if get_col(1) is None:
            cur_station = get_col(0)
        else:
            portion, err = parse_portion(get_col(4))

            station_meal, station_name = cur_station.split(' - ', 1)
            station_meal = station_meal.lower()

            if station_name.lower() not in ['homestyle', 'rooted', 'fyul', 'flame', '500 degrees', 'carved and crafted']:
                err = True
            else:
                if err:
                    unparseable += 1

            if not err:
                name = re.sub(r'(CHE( (\d+))? (- )?)|(HC )|([\w+]+: )',
                               '', get_col('A'), count=1)[:MAX_NAME_LEN]
                meal_items[get_col(0)] = {
                    # basic info
                    'name': name,
                    'school': SCHOOL_ID,
                    'version': version,
                    'station': station_name,

                    # basic nutritional/number info
                    'category': c.lower() if (c := get_col(3)) else None,
                    'cafeteria_id': get_col('B'),
                    'portion_weight': num_col('H'),
                    'portion_volume': portion,
                    'ingredients': [],

                    # sheet 1 nutrients
                    'calories': num_col('I'),
                    'protein': num_col('K'),
                    'total_fat': num_col('L'),
                    'carbohydrate': num_col('M'),
                    'fiber': num_col('N'),
                    'saturated_fat': num_col('O'),
                    'potassium': num_col('Q'),
                    'sodium': num_col('T'),
                    'sugar': num_col('W'),

                    # Extra property for other parsers
                    'meal': station_meal,
                }

    for row in micros_sheet.iter_rows(13):
        def get_col(col: Union[int, str]):
            if isinstance(col, int):
                return row[col].value
            elif isinstance(col, str) and len(col) == 1:
                return row[ord(col) - 65].value
            else:
                raise ValueError(f'Unreadable column {col}')

        def num_col(col: Union[int, str]):
            val = get_col(col)
            if val is None:
                return 0
            else:
                only_num = re.sub(r'[^0-9.]', '', val)
                return float(only_num) if only_num else 0

        if get_col(0) in meal_items:
            meal_items[get_col(0)] |= {
                'cholesterol': num_col('H'),
                'calcium': num_col('J'),
                'iron': num_col('K'),
                'vitamin_d': num_col('L'),
                'vitamin_c': num_col('O'),
                'vitamin_a': num_col('P'),
            }

    print(f'Parsed {len(meal_items)} meal items')
    print(f'Got {unparseable} items with correct stations but otherwise unreadable')

    return meal_items, bad_units


def add_meal_items(version):
    meal_items, _ = parse_meal_items(version)

    class MealItemSerializer(serializers.ModelSerializer):
        class Meta:
            model = MealItem
            exclude = ['ingredients']

    meals = []
    trim_ids = []
    objs = []
    for m_info in meal_items.values():
        ser = MealItemSerializer(data=m_info)
        ser.is_valid(raise_exception=True)
        objs.append(MealItem(**ser.validated_data))
        meals.append(str(m_info['meal']))
        trim_ids.append(str(m_info['cafeteria_id']))

    return list(zip(MealItem.objects.bulk_create(objs), meals, trim_ids))


def main():
    meal_items, bad_units = parse_meal_items(-1)

    with open(OUT_FILE_PATH, 'w') as f:
        json.dump(meal_items, f)

    print('-- Bad Units --')
    for unit in sorted(bad_units):
        print(unit)


if __name__ == '__main__':
    main()
